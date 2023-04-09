from openpyxl import load_workbook
from openpyxl.styles import Alignment

wb = load_workbook(filename='function.xlsx')
ws = wb['Sheet1']

# 셀 병합
# ws.merge_cells('A8:E10')
ws.cell(column=1,row=8, value="제목")

# 셀 병합 후, 병합된 셀 내부의 텍스트를 중앙에 위치시키기 위해 Alignment 객체 생성
align = Alignment(horizontal='center' , vertical='center')

# 병합된 셀에 스타일 지정
ws['A8'].alignment = align

wb.save('function.xlsx')
wb.close()


# for j in range(2,6):
#     for i in range(1,11):
#         ws.cell(column=j,row=i, value=(ws.cell(column=j-1,row=i).value+3))

# wb.save('function.xlsx')
# wb.close()



# for row in ws.rows:
#     for cell in row:
#         print(cell.value)
#

